# -*- coding: UTF-8 -*-
import os
import xlrd
import datetime
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
from mysql import connector


def readAllInfo():
    workbook = xlrd.open_workbook('result/TOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = []
    for i in range(1, sheet.nrows):
        line = sheet.row_values(i)
        if len(line[0]) == 0:
            continue
        if isinstance(sheet.cell_value(i, 6), float):
            tmpTuple = xlrd.xldate_as_tuple(line[6], 0)
            line[6] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        if isinstance(sheet.cell_value(i, 7), float):
            tmpTuple = xlrd.xldate_as_tuple(line[7], 0)
            line[7] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        for j in range(0, len(line)):
            if isinstance(line[j], str) == False:
                line[j] = str(line[j])
        lineInfos.append(line)
    return lineInfos


def transTopNInfoToCSV():
    workbook = xlrd.open_workbook('result/TOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = ""
    for i in range(0, sheet.nrows):
        line = sheet.row_values(i)
        if i != 0:
            if isinstance(sheet.cell_value(i, 6), float):
                tmpTuple = xlrd.xldate_as_tuple(line[6], 0)
                line[6] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

            if isinstance(sheet.cell_value(i, 7), float):
                tmpTuple = xlrd.xldate_as_tuple(line[7], 0)
                line[7] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

            for j in range(0, len(line)):
                if isinstance(line[j], str) == False:
                    line[j] = str(line[j])
        for val in line:
            lineInfos += "%s," % (str(val).replace('\n', ';').replace('\r', ';').replace(',', ';'))
        lineInfos += '\n'
    return lineInfos


def readRatioAllInfo():
    workbook = xlrd.open_workbook('result/双流比TOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = []
    for i in range(1, sheet.nrows):
        line = sheet.row_values(i)
        if len(line[0]) == 0:
            continue
        if isinstance(sheet.cell_value(i, 6), float):
            tmpTuple = xlrd.xldate_as_tuple(line[6], 0)
            line[6] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        if isinstance(sheet.cell_value(i, 7), float):
            tmpTuple = xlrd.xldate_as_tuple(line[7], 0)
            line[7] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        for j in range(0, len(line)):
            if isinstance(line[j], str) == False:
                line[j] = str(line[j])
        lineInfos.append(line)
    return lineInfos


def trans4GTopNInfoToCSV():
    workbook = xlrd.open_workbook('result/双流比TOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = ""
    for i in range(0, sheet.nrows):
        line = sheet.row_values(i)
        if i != 0:
            if isinstance(sheet.cell_value(i, 6), float):
                tmpTuple = xlrd.xldate_as_tuple(line[6], 0)
                line[6] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

            if isinstance(sheet.cell_value(i, 7), float):
                tmpTuple = xlrd.xldate_as_tuple(line[7], 0)
                line[7] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

            for j in range(0, len(line)):
                if isinstance(line[j], str) == False:
                    line[j] = str(line[j])
        for val in line:
            lineInfos += "%s," % (str(val).replace('\n', ';').replace('\r', ';').replace(',', ';'))
        lineInfos += '\n'
    return lineInfos


def readRatioSerchInfo(carrier):
    workbook = xlrd.open_workbook('result/双流比TOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = []
    for i in range(1, sheet.nrows):
        line = sheet.row_values(i)
        if len(line[0]) == 0 or line[0] != carrier:
            continue

        if isinstance(sheet.cell_value(i, 6), float):
            tmpTuple = xlrd.xldate_as_tuple(line[6], 0)
            line[6] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        if isinstance(sheet.cell_value(i, 7), float):
            tmpTuple = xlrd.xldate_as_tuple(line[7], 0)
            line[7] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        for j in range(0, len(line)):
            if isinstance(line[j], str) == False:
                line[j] = str(line[j])
        lineInfos.append(line)
    return lineInfos


def readRccAllInfo():
    workbook = xlrd.open_workbook('result/RRCTOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = []
    for i in range(1, sheet.nrows):
        line = sheet.row_values(i)
        if len(line[0]) == 0:
            continue

        if isinstance(sheet.cell_value(i, 7), float):
            tmpTuple = xlrd.xldate_as_tuple(line[7], 0)
            line[7] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        if isinstance(sheet.cell_value(i, 8), float):
            tmpTuple = xlrd.xldate_as_tuple(line[8], 0)
            line[8] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        for j in range(0, len(line)):
            if isinstance(line[j], str) == False:
                line[j] = str(line[j])
        lineInfos.append(line)
    return lineInfos


def transRccInfoToCSV():
    workbook = xlrd.open_workbook('result/RRCTOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = ""
    for i in range(0, sheet.nrows):
        line = sheet.row_values(i)
        if i != 0:
            if isinstance(sheet.cell_value(i, 7), float):
                tmpTuple = xlrd.xldate_as_tuple(line[7], 0)
                line[7] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

            if isinstance(sheet.cell_value(i, 8), float):
                tmpTuple = xlrd.xldate_as_tuple(line[8], 0)
                line[8] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

            for j in range(0, len(line)):
                if isinstance(line[j], str) == False:
                    line[j] = str(line[j])
        for val in line:
            lineInfos += "%s," % (str(val).replace('\n', ';').replace('\r', ';').replace(',', ';'))
        lineInfos += '\n'
    return lineInfos


def readCarrierInfo(carrier):
    workbook = xlrd.open_workbook('result/TOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = []
    for i in range(1, sheet.nrows):
        line = sheet.row_values(i)
        if len(line[0]) == 0 or line[0] != carrier:
            continue

        if isinstance(sheet.cell_value(i, 6), float):
            tmpTuple = xlrd.xldate_as_tuple(line[6], 0)
            line[6] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        if isinstance(sheet.cell_value(i, 7), float):
            tmpTuple = xlrd.xldate_as_tuple(line[7], 0)
            line[7] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        for j in range(0, len(line)):
            if isinstance(line[j], str) == False:
                line[j] = str(line[j])
        lineInfos.append(line)
    return lineInfos


def readInterfereInfo():
    workbook = xlrd.open_workbook('result/干扰源信息汇总.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = []
    for i in range(1, sheet.nrows):
        line = sheet.row_values(i)
        if isinstance(sheet.cell_value(i, 9), float):
            tmpTuple = xlrd.xldate_as_tuple(line[9], 0)
            line[9] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

        for j in range(0, len(line)):
            if isinstance(line[j], float) and j != 5 and j != 6:
                line[j] = str(int(line[j]))
        lineInfos.append(line)

    return lineInfos


def transInterfereInfoToCSV():
    workbook = xlrd.open_workbook('result/干扰源信息汇总.xlsx')
    sheet = workbook.sheets()[0]

    lineInfos = ""
    for i in range(0, sheet.nrows):
        line = sheet.row_values(i)
        if i != 0:
            if isinstance(sheet.cell_value(i, 9), float):
                tmpTuple = xlrd.xldate_as_tuple(line[9], 0)
                line[9] = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])

            for j in range(0, len(line)):
                if isinstance(line[j], float) and j != 5 and j != 6:
                    line[j] = str(int(line[j]))
        for val in line:
            lineInfos += "%s," % (str(val).replace('\n', ';').replace('\r', ';').replace(',', ';'))
        lineInfos += '\n'
    return lineInfos


def getCountInfo(sdate, edate):
    workbook = xlrd.open_workbook('result/TOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]

    recordCount = 0
    closedCount = 0
    leavedCount = 0
    chargeDict = {'长信泰康': 0, '宜通': 0, '省工程': 0}
    reasonDict = dict()

    for i in range(1, sheet.nrows):
        date_str = ""
        if isinstance(sheet.cell_value(i, 6), float):
            tmpTuple = xlrd.xldate_as_tuple(sheet.cell_value(i, 6), 0)
            date_str = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])
        elif isinstance(sheet.cell_value(i, 6), str):
            try:
                date_str = sheet.cell_value(i, 6)
                tmp_date = datetime.datetime.strptime(date_str, '%Y/%m/%d')
            except Exception:
                date_str = ""

        if len(date_str) != 0:
            tmp_date = datetime.datetime.strptime(date_str, '%Y/%m/%d')
            if sdate <= tmp_date and tmp_date <= edate:
                recordCount = recordCount + 1
                if isinstance(sheet.cell_value(i, 7), float):
                    closedCount = closedCount + 1
                elif isinstance(sheet.cell_value(i, 7), str) and len(sheet.cell_value(i, 7)) > 0:
                    closedCount = closedCount + 1
                else:
                    leavedCount = leavedCount + 1
                chargeKey = sheet.cell_value(i, 2)
                if chargeKey in chargeDict:
                    chargeDict[chargeKey] = chargeDict[chargeKey] + 1
                else:
                    chargeDict[chargeKey] = 1

                reasonKey = sheet.cell_value(i, 3).split('、')[0]
                if reasonKey in reasonDict:
                    reasonDict[reasonKey] = reasonDict[reasonKey] + 1
                else:
                    reasonDict[reasonKey] = 1

    countInfo = []
    countInfo.append(recordCount)
    countInfo.append(closedCount)
    countInfo.append(leavedCount)
    countInfo.append(chargeDict)
    countInfo.append(reasonDict)

    return countInfo


clr = ["Red", "Blue", "ForestGreen", "Black", "LawnGreen", "Purple", "Green", "Violet", "Sienna", "GoldenRod"]


def getRatioTopN(sdate, edate):
    filename = 'result/' + sdate + '_' + edate + '_CQIResult.xlsx'
    datas = list()
    if os.path.exists(filename) == False:
        return datas

    workbook = xlrd.open_workbook(filename)
    sheet = workbook.sheets()[0]

    for i in range(1, sheet.nrows):
        datas.append(sheet.row_values(i))

    return datas


def getRatioCountInfo(sdate, edate):
    workbook = xlrd.open_workbook('result/双流比TOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]
    recordCount = 0
    closedCount = 0
    leavedCount = 0
    chargeDict = {}
    reasonDict = dict()

    for i in range(1, sheet.nrows):
        date_str = ""
        if isinstance(sheet.cell_value(i, 6), float):
            tmpTuple = xlrd.xldate_as_tuple(sheet.cell_value(i, 6), 0)
            date_str = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])
        elif isinstance(sheet.cell_value(i, 6), str):
            try:
                date_str = sheet.cell_value(i, 6)
                tmp_date = datetime.datetime.strptime(date_str, '%Y/%m/%d')
            except Exception:
                date_str = ""

        if len(date_str) != 0:
            tmp_date = datetime.datetime.strptime(date_str, '%Y/%m/%d')
            if sdate <= tmp_date and tmp_date <= edate:
                recordCount = recordCount + 1
                if isinstance(sheet.cell_value(i, 7), float):
                    closedCount = closedCount + 1
                elif isinstance(sheet.cell_value(i, 7), str) and len(sheet.cell_value(i, 7)) > 0:
                    closedCount = closedCount + 1
                else:
                    leavedCount = leavedCount + 1
                chargeKey = sheet.cell_value(i, 2)
                if chargeKey in chargeDict:
                    chargeDict[chargeKey][0] = chargeDict[chargeKey][0] + 1
                else:
                    chargeDict[chargeKey] = [1, clr[len(chargeDict)]]

                reasonKey = sheet.cell_value(i, 3).split('、')[0]
                if reasonKey in reasonDict:
                    reasonDict[reasonKey] = reasonDict[reasonKey] + 1
                else:
                    reasonDict[reasonKey] = 1

    countInfo = []
    countInfo.append(recordCount)
    countInfo.append(closedCount)
    countInfo.append(leavedCount)
    countInfo.append(chargeDict)
    countInfo.append(reasonDict)

    return countInfo


def getRccCountInfo(sdate, edate):
    workbook = xlrd.open_workbook('result/RRCTOPN问题解决统计.xlsx')
    sheet = workbook.sheets()[0]
    recordCount = 0
    closedCount = 0
    leavedCount = 0
    chargeDict = {}
    reasonDict = dict()

    for i in range(1, sheet.nrows):
        date_str = ""
        if isinstance(sheet.cell_value(i, 7), float):
            tmpTuple = xlrd.xldate_as_tuple(sheet.cell_value(i, 7), 0)
            date_str = str(tmpTuple[0]) + '/' + str(tmpTuple[1]) + '/' + str(tmpTuple[2])
        elif isinstance(sheet.cell_value(i, 7), str):
            try:
                date_str = sheet.cell_value(i, 7)
                tmp_date = datetime.datetime.strptime(date_str, '%Y/%m/%d')
            except Exception:
                date_str = ""

        if len(date_str) != 0:
            tmp_date = datetime.datetime.strptime(date_str, '%Y/%m/%d')
            if sdate <= tmp_date and tmp_date <= edate:
                recordCount = recordCount + 1
                if isinstance(sheet.cell_value(i, 8), float):
                    closedCount = closedCount + 1
                elif isinstance(sheet.cell_value(i, 8), str) and len(sheet.cell_value(i, 8)) > 0:
                    closedCount = closedCount + 1
                else:
                    leavedCount = leavedCount + 1
                chargeKey = sheet.cell_value(i, 2)
                if chargeKey in chargeDict:
                    chargeDict[chargeKey][0] = chargeDict[chargeKey][0] + 1
                else:
                    chargeDict[chargeKey] = [1, clr[len(chargeDict)]]

                reasonKey = sheet.cell_value(i, 4).split('、')[0]
                if reasonKey in reasonDict:
                    reasonDict[reasonKey] = reasonDict[reasonKey] + 1
                else:
                    reasonDict[reasonKey] = 1

    countInfo = []
    countInfo.append(recordCount)
    countInfo.append(closedCount)
    countInfo.append(leavedCount)
    countInfo.append(chargeDict)
    countInfo.append(reasonDict)

    return countInfo


def get_eNodeBName():
    res = dict()
    workbook = xlrd.open_workbook('佛山无线中心LTE工参.xlsx')
    sheet = workbook.sheets()[1]
    for r in range(1, sheet.nrows):
        res[sheet.cell_value(r, 9)] = sheet.cell_value(r, 10)
    return res


def getRRCTopN(date):
    defVal = ['', '', 0, 0, 0, 0]
    data = dict()
    eNodeNames = get_eNodeBName()
    if os.path.exists("sqlop/data/%s_rrc_d.txt"%(date.strftime('%Y%m%d'))) is False:
        sqlStr = 'select * from tle_topn.rrc_conn_h where BEGINCOLLECTTIME >= \'%s 00:00\' and BEGINCOLLECTTIME <= \'%s 23:59\'' % (
        date.strftime('%Y-%m-%d'), date.strftime('%Y-%m-%d'))
        cnx = connector.connect(host='219.128.254.36', user='root', password='Abcdef9*')
        cursor = cnx.cursor()
        cursor.execute(sqlStr)
        res = cursor.fetchall()

        for r in res:
            cellId = '%d_%d' % (r[1], r[2])
            if cellId not in data:
                data[cellId] = defVal.copy()
                data[cellId][0] = cellId
                if cellId in eNodeNames:
                    data[cellId][1] = eNodeNames[cellId]
            data[cellId][2] += r[3]
            data[cellId][3] += r[4]

        with open('sqlop/data/%s_rrc_d.txt'%(date.strftime('%Y%m%d')), 'w') as ff:
            for key, val in data.items():
                ff.write('%s,%d,%d\n'%(key, val[2], val[3]))
            ff.close()
    else:
        with open("sqlop/data/%s_rrc_d.txt"%(date.strftime('%Y%m%d')), 'r') as ff:
            for line in ff.readlines():
                line = line.replace('\r', '')
                line = line.replace('\n', '')
                line = line.split(',')
                cellId = line[0]
                if cellId not in data:
                    data[cellId] = defVal.copy()
                    data[cellId][0] = cellId
                    if cellId in eNodeNames:
                        data[cellId][1] = eNodeNames[cellId]
                data[cellId][2] += int(line[1])
                data[cellId][3] += int(line[2])

    dataList = list()
    for k, v in data.items():
        dataList.append(v)
    for d in dataList:
        if d[3]:
            d[4] = d[2] / d[3] * 100
    dataList.sort(key=lambda x: x[4] - x[3], reverse=True)
    nCnt = len(dataList)
    s = 10 if nCnt >= 10 else nCnt
    resData = list()
    for d in dataList:
        if s == 0:
            break
        if d[3] == 0:
            continue
        resData.append(d)
        s = s - 1
    return resData


def updateRowInfo(filename, row, rowData):
    rowTxt = str(row + 1)
    wb = load_workbook(filename)
    ws = wb.active

    colCount = len(rowData) + 1
    for i in range(1, colCount):
        cellLetter = get_column_letter(i) + rowTxt
        print(cellLetter)
        ws[cellLetter] = rowData[i - 1]

    wb.save(filename)


def getRowIdx(filename, idx, carrier):
    workbook = xlrd.open_workbook(filename)
    sheet = workbook.sheets()[0]

    for i in range(1, sheet.nrows):
        if sheet.cell_value(i, 0) == carrier:
            idx -= 1
            if idx == 0:
                return i

    return sheet.nrows
